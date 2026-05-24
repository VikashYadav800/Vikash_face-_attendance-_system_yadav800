"""
 Face Recognition — Unified Backend
All services merged into one FastAPI app running on port 8000.

Endpoints:
  Face Recognition (original app.py):
    POST /build-db
    GET  /list-db
    GET  /search/{username}
    POST /match-image

  User CRUD (original crud.py):
    POST   /add-user
    DELETE /delete-user/{name}
    GET    /users
    GET    /search-user/{name}

  Dashboard + Excel (original excel.py):
    GET /dashboard-stats
    GET /attendance-by-date?date=YYYY-MM-DD
    GET /yearly-attendance?year=2026

Run:
    uvicorn server:app --host 0.0.0.0 --port 8000 --reload
"""

import os
import shutil
import hashlib
import calendar
import requests
import numpy as np
import cv2
import faiss
import mysql.connector

from datetime import datetime, timedelta
from fastapi import FastAPI, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from insightface.app import FaceAnalysis
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────



DATASET_DIR        = "dataset"
FAISS_DIM          = 512
MATCH_THRESHOLD    = 0.6          # cosine similarity threshold for face match
SIMILARITY_THRESH  = 0.92         # duplicate-face guard in add-user

DB_CONFIG = dict(
    host     = "localhost",
    user     = "root",
    password = "root",
    database = "face_recognition",
)

# ─────────────────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────────────────

app = FastAPI(title="Face Unified API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────────────────
# DATABASE HELPERS
# ─────────────────────────────────────────────────────────

def get_db():
    """Return a fresh MySQL connection (caller must close it)."""
    return mysql.connector.connect(**DB_CONFIG)


# ─────────────────────────────────────────────────────────
# INSIGHTFACE — single model instance shared across routes
# ─────────────────────────────────────────────────────────

face_app = FaceAnalysis(name="buffalo_l")
face_app.prepare(ctx_id=-1, det_size=(640, 640))

# ─────────────────────────────────────────────────────────
# FAISS INDEX — loaded once at startup
# ─────────────────────────────────────────────────────────

faiss_index: faiss.IndexFlatIP = faiss.IndexFlatIP(FAISS_DIM)
face_names: list[str] = []


def _normalize(v: np.ndarray) -> np.ndarray:
    norm = np.linalg.norm(v)
    if norm == 0:
        return v
    return v / norm


def _load_faiss_from_db() -> None:
    """Populate the in-memory FAISS index from the faces table.

    Rows whose blobs don't match FAISS_DIM are skipped so a single
    corrupt/stale record never crashes startup. Run POST /build-db to fix.
    """
    global faiss_index, face_names
    faiss_index = faiss.IndexFlatIP(FAISS_DIM)
    face_names  = []

    conn   = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT name, embedding FROM faces")
        rows = cursor.fetchall()
        if not rows:
            print("[FAISS] faces table is empty — call POST /build-db or POST /add-user.")
            return

        embeddings, skipped = [], 0
        for name, blob in rows:
            emb = np.frombuffer(blob, dtype=np.float32).copy()
            if emb.shape[0] != FAISS_DIM:
                print(f"[FAISS] WARNING: skipping '{name}' — "
                      f"{emb.shape[0]} dims (expected {FAISS_DIM}). "
                      "Run POST /build-db to rebuild.")
                skipped += 1
                continue
            emb = _normalize(emb)
            embeddings.append(emb)
            face_names.append(name)

        if embeddings:
            faiss_index.add(np.array(embeddings, dtype="float32"))

        msg = f"[FAISS] Loaded {len(embeddings)} embeddings"
        msg += f", skipped {skipped} bad rows." if skipped else "."
        print(msg)
    finally:
        cursor.close()
        conn.close()


_load_faiss_from_db()

# ─────────────────────────────────────────────────────────
# EXCEL STYLES (from excel.py — unchanged)
# ─────────────────────────────────────────────────────────

name_bg_fill          = PatternFill(start_color="2C3E6B", end_color="2C3E6B", fill_type="solid")
sno_bg_fill           = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
present_fill          = PatternFill(start_color="D6F0E0", end_color="D6F0E0", fill_type="solid")
absent_fill           = PatternFill(start_color="FAD7D7", end_color="FAD7D7", fill_type="solid")
sunday_fill           = PatternFill(start_color="D8DCE6", end_color="D8DCE6", fill_type="solid")
checkout_missing_fill = PatternFill(start_color="FDE8C8", end_color="FDE8C8", fill_type="solid")
festival_fill         = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")

month_end_fills = [
    PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
    PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
    PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
    PatternFill(start_color="C5E1F5", end_color="C5E1F5", fill_type="solid"),
    PatternFill(start_color="F8D7E8", end_color="F8D7E8", fill_type="solid"),
]

white_font    = Font(color="FFFFFF", bold=True,  name="Calibri", size=10)
header_font   = Font(bold=True, color="FFFFFF",  name="Calibri", size=11)
name_font     = Font(bold=True, color="FFFFFF",  name="Calibri", size=11)
dark_font     = Font(color="2C3E6B", bold=True,  name="Calibri", size=10)
present_font  = Font(color="1A6B3A", bold=True,  name="Calibri", size=9)
absent_font   = Font(color="9B1C1C", bold=True,  name="Calibri", size=10)
sunday_font   = Font(color="5A6175", bold=True,  name="Calibri", size=10)
checkout_font = Font(color="7A4000", bold=True,  name="Calibri", size=9)
festival_font = Font(color="5B2D8E", bold=True,  name="Calibri", size=10)
summary_font  = Font(color="1A2E5A", bold=True,  name="Calibri", size=10)

thin_side   = Side(style="thin",   color="B0BAD0")
med_side    = Side(style="medium", color="8B9CC0")
cell_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border  = Border(left=med_side,  right=med_side,  top=med_side,  bottom=med_side)
checkout_border = Border(
    left=Side(style="medium",  color="E07B00"),
    right=Side(style="medium", color="E07B00"),
    top=Side(style="medium",   color="E07B00"),
    bottom=Side(style="medium",color="E07B00"),
)

FESTIVAL_DAYS = ["2026-01-26", "2026-08-15", "2026-10-02"]

MONTH_END_COLUMNS = [
    "Paid Leave of Month",
    "Actual Leave taken",
    "Actual Deduction",
    "Paid Leave Balance",
    "Paid Leave to be Carry Forward",
    "Total day = 31",
    "In Timings Average",
    "Total In",
    "Total Out",
]

_TYPE_FONT = {
    "present":          present_font,
    "absent":           absent_font,
    "checkout_missing": checkout_font,
    "sunday":           sunday_font,
    "festival":         festival_font,
    "leave":            festival_font,
}


def _is_off_day(year: int, month: int, day: int) -> bool:
    weekday = datetime(year, month, day).weekday()
    if weekday == 6:
        return True
    if weekday == 5:
        week_of_month = (day - 1) // 7 + 1
        if week_of_month in [2, 4]:
            return True
    return False


def _apply_excel_styling(ws, row_number: int, cell_data: list, days_in_month: int, name: str) -> None:
    ctr      = Alignment(horizontal="center", vertical="center")
    ctr_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=row_number, column=1).fill      = sno_bg_fill
    ws.cell(row=row_number, column=1).font      = dark_font
    ws.cell(row=row_number, column=1).alignment = ctr
    ws.cell(row=row_number, column=1).border    = cell_border

    nc = ws.cell(row=row_number, column=2)
    nc.fill      = name_bg_fill
    nc.font      = name_font
    nc.alignment = ctr_wrap
    nc.border    = cell_border

    month_end_start = 3 + days_in_month
    for i in range(9):
        c = ws.cell(row=row_number, column=month_end_start + i)
        c.fill      = month_end_fills[i]
        c.font      = summary_font
        c.alignment = ctr
        c.border    = cell_border

    for col_index, info in enumerate(cell_data, start=3):
        if info["type"] == "blank":
            continue
        c = ws.cell(row=row_number, column=col_index)
        c.fill      = info["fill"]
        c.font      = _TYPE_FONT.get(info["type"], dark_font)
        c.alignment = ctr_wrap
        c.border    = checkout_border if info["type"] == "checkout_missing" else cell_border


# ─────────────────────────────────────────────────────────
# ── FACE RECOGNITION ROUTES (app.py) ──────────────────────
# ─────────────────────────────────────────────────────────

@app.post("/build-db")
def build_database():
    """Re-build the FAISS index and faces table from the dataset folder."""
    global faiss_index, face_names

    conn   = get_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM faces")
    conn.commit()

    faiss_index = faiss.IndexFlatIP(FAISS_DIM)
    face_names  = []
    total       = 0

    for person_name in os.listdir(DATASET_DIR):
        person_path = os.path.join(DATASET_DIR, person_name)
        if not os.path.isdir(person_path):
            continue

        for img_name in os.listdir(person_path):
            img_path = os.path.join(person_path, img_name)
            img      = cv2.imread(img_path)
            if img is None:
                continue

            faces = face_app.get(img)
            if not faces:
                continue

            emb = _normalize(faces[0].embedding.astype("float32"))
            faiss_index.add(np.array([emb]))
            face_names.append(person_name)

            cursor.execute(
                "INSERT INTO faces (name, embedding) VALUES (%s,%s)",
                (person_name, emb.tobytes()),
            )
            total += 1

    conn.commit()
    cursor.close()
    conn.close()

    return {"status": "success", "total_embeddings": total}


@app.get("/list-db")
def list_database():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM faces")
    data = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]
    cursor.close()
    conn.close()
    return {"total": len(data), "data": data}


@app.get("/search/{username}")
def search_face(username: str):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM faces WHERE name=%s", (username,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    if not rows:
        return {"match": False}
    return {"match": True, "total": len(rows), "data": rows}


@app.post("/match-image")
async def match_image(file: UploadFile = File(...)):
    contents = await file.read()
    nparr = np.frombuffer(contents, np.uint8)
    img   = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    if img is None:
        return {"match": False}

    faces = face_app.get(img)
    if not faces:
        return {"match": False}

    emb = _normalize(faces[0].embedding.astype("float32")).reshape(1, -1)
    D, I = faiss_index.search(emb, 1)
    score   = float(D[0][0])
    best_id = int(I[0][0])

    if score < MATCH_THRESHOLD:
        return {"match": False}

    name = face_names[best_id]
    now  = datetime.now()
    today = now.date()

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT check_in, check_out FROM attendance WHERE person_name=%s AND date=%s",
        (name, today),
    )
    record = cursor.fetchone()

    if record is None:
        cursor.execute(
            "INSERT INTO attendance (person_name, date, check_in, status) VALUES (%s,%s,%s,%s)",
            (name, today, now.strftime("%Y-%m-%d %H:%M:%S"), "Checked In"),
        )
        conn.commit()
        cursor.close(); conn.close()
        return {"match": True, "name": name, "attendance": "Check-in marked"}

    checkin_time, checkout_time = record

    if checkout_time:
        cursor.close(); conn.close()
        return {"match": True, "name": name, "attendance": "Attendance already completed"}

    # check_in is a DATETIME column — MySQL connector returns a datetime object directly
    if isinstance(checkin_time, datetime):
        checkin_dt = checkin_time
    elif hasattr(checkin_time, "total_seconds"):
        # Fallback: TIME column returned as timedelta
        total_secs = int(checkin_time.total_seconds())
        h, rem = divmod(total_secs, 3600)
        m, s   = divmod(rem, 60)
        from datetime import time as dt_time
        checkin_dt = datetime.combine(today, dt_time(h, m, s))
    else:
        checkin_dt = datetime.combine(today, checkin_time)

    if now < checkin_dt + timedelta(hours=1):
        cursor.close(); conn.close()
        return {"match": True, "name": name, "attendance": "Checkout allowed after 1 hour"}

    cursor.execute(
        "UPDATE attendance SET check_out=%s, status=%s WHERE person_name=%s AND date=%s",
        (now.strftime("%Y-%m-%d %H:%M:%S"), "Completed", name, today),
    )
    conn.commit()
    cursor.close(); conn.close()
    return {"match": True, "name": name, "attendance": "Checkout marked"}


# ─────────────────────────────────────────────────────────
# ── USER CRUD ROUTES (crud.py) ─────────────────────────────
# ─────────────────────────────────────────────────────────

@app.post("/add-user")
async def add_user(name: str = Form(...), image: UploadFile = File(...)):
    global faiss_index, face_names

    conn   = get_db()
    cursor = conn.cursor()

    try:
        contents = await image.read()
        if not contents:
            return {"status": "error", "message": "Empty image"}

        # ───────────── IMAGE HASH CHECK ─────────────
        image_hash = hashlib.md5(contents).hexdigest()
        person_dir = os.path.join(DATASET_DIR, name)
        os.makedirs(person_dir, exist_ok=True)

        for fname in os.listdir(person_dir):
            fpath = os.path.join(person_dir, fname)
            with open(fpath, "rb") as f:
                if hashlib.md5(f.read()).hexdigest() == image_hash:
                    return {"status": "duplicate", "message": "Image already exists"}

        # ───────────── FACE DETECTION ─────────────
        npimg = np.frombuffer(contents, np.uint8)
        img   = cv2.imdecode(npimg, cv2.IMREAD_COLOR)

        if img is None:
            return {"status": "error", "message": "Invalid image"}

        faces = face_app.get(img)   # ✅ use SAME model everywhere

        if len(faces) == 0:
            return {"status": "error", "message": "No face detected"}
        if len(faces) > 1:
            return {"status": "error", "message": "Multiple faces detected"}

        # ───────────── EMBEDDING ─────────────
        emb = _normalize(faces[0].embedding.astype("float32"))

        # ───────────── DUPLICATE FACE CHECK ─────────────
        cursor.execute("SELECT name, embedding FROM faces")
        for db_name, blob in cursor.fetchall():
            stored_emb = np.frombuffer(blob, dtype=np.float32)

            sim = np.dot(emb, stored_emb)
            if sim > SIMILARITY_THRESH:
                return {
                    "status": "duplicate",
                    "message": f"Already registered as {db_name}"
                }

        # ───────────── SAVE IMAGE ─────────────
        filename = image.filename or f"{name}.jpg"
        with open(os.path.join(person_dir, filename), "wb") as f:
            f.write(contents)

        # ───────────── SAVE TO DB (BLOB) ─────────────
        cursor.execute(
            "INSERT INTO faces (name, embedding) VALUES (%s,%s)",
            (name, emb.tobytes())
        )
        conn.commit()

        # ───────────── UPDATE FAISS (NO FULL REBUILD) ─────────────
        faiss_index.add(np.array([emb], dtype="float32"))
        face_names.append(name)

        return {"status": "success", "message": "User added"}

    except Exception as e:
        return {"status": "error", "message": str(e)}

    finally:
        cursor.close()
        conn.close()
        
@app.delete("/delete-user/{name}")
def delete_user(name: str):
    conn   = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT COUNT(*) as count FROM faces WHERE name=%s", (name,))
        if cursor.fetchone()["count"] == 0:
            return {"status": "error", "message": f"No member '{name}'"}

        folder = os.path.join(DATASET_DIR, name)
        if os.path.exists(folder):
            shutil.rmtree(folder)

        cursor.execute("DELETE FROM faces WHERE name=%s", (name,))
        conn.commit()

        # Also remove from attendance (optional — comment out if not desired)
        # cursor.execute("DELETE FROM attendance WHERE person_name=%s", (name,))
        # conn.commit()

        _rebuild_faiss_internal(cursor)

        return {"status": "success", "message": "Deleted"}

    except Exception as e:
        return {"status": "error", "message": str(e)}

    finally:
        cursor.close()
        conn.close()


@app.get("/users")
def list_users():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT name FROM faces")
    users = [r[0] for r in cursor.fetchall()]
    cursor.close(); conn.close()
    return {"total": len(users), "users": users}


@app.get("/search-user/{name}")
def search_user(name: str):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM faces WHERE name=%s", (name,))
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    if not rows:
        return {"found": False}
    return {"found": True, "data": rows}


def _rebuild_faiss_internal(existing_cursor) -> None:
    """Re-populate FAISS from DB using an already-open cursor."""
    global faiss_index, face_names
    faiss_index = faiss.IndexFlatIP(FAISS_DIM)
    face_names  = []
    existing_cursor.execute("SELECT name, embedding FROM faces")
    rows = existing_cursor.fetchall()
    embeddings = []
    for row in rows:
        name  = row[0] if isinstance(row, tuple) else row["name"]
        blob  = row[1] if isinstance(row, tuple) else row["embedding"]
        try:
            emb = np.frombuffer(blob, dtype=np.float32).copy()
        except Exception:
            emb = np.array([float(x) for x in blob.split(",")], dtype=np.float32)
        if emb.shape[0] != FAISS_DIM:
            print(f"[FAISS] _rebuild: skipping '{name}' — wrong dim {emb.shape[0]}")
            continue
        emb = _normalize(emb)
        embeddings.append(emb)
        face_names.append(name)
    if embeddings:
        faiss_index.add(np.array(embeddings, dtype="float32"))


# ─────────────────────────────────────────────────────────
# ── DASHBOARD + EXCEL ROUTES (excel.py) ───────────────────
# ─────────────────────────────────────────────────────────

@app.get("/dashboard-stats")
def dashboard_stats():
    conn   = get_db()
    cursor = conn.cursor(dictionary=True)
    today  = datetime.today().date()

    cursor.execute("SELECT COUNT(DISTINCT name) as total FROM faces")
    employees = cursor.fetchone()["total"]

    cursor.execute(
        "SELECT COUNT(DISTINCT person_name) as present FROM attendance WHERE DATE(date)=%s AND check_in IS NOT NULL",
        (today,),
    )
    present = cursor.fetchone()["present"]
    absent  = employees - present

    cursor.execute(
        "SELECT COUNT(DISTINCT person_name) as late FROM attendance WHERE DATE(date)=%s AND check_in IS NOT NULL AND TIME(check_in) > '09:30:00'",
        (today,),
    )
    late = cursor.fetchone()["late"]

    weekly = []
    for i in range(29, -1, -1):
        day_date = today - timedelta(days=i)
        cursor.execute(
            "SELECT COUNT(DISTINCT person_name) as present FROM attendance WHERE DATE(date)=%s AND check_in IS NOT NULL",
            (day_date,),
        )
        p = cursor.fetchone()["present"] or 0
        weekly.append({"day": day_date.strftime("%d %b"), "present": p, "absent": employees - p})

    cursor.close(); conn.close()
    return {"stats": {"employees": employees, "present": present, "absent": absent, "late": late}, "weekly": weekly}


@app.get("/attendance-by-date")
def attendance_by_date(date: str = Query(...)):
    conn   = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, person_name, date, check_in, check_out, status FROM attendance WHERE date=%s",
        (date,),
    )
    rows   = cursor.fetchall()
    result = []
    for r in rows:
        if r["check_in"] and r["check_out"]:
            diff  = r["check_out"] - r["check_in"]
            hours = diff.total_seconds() / 3600
            wh    = f"{hours:.2f} hrs"
        elif r["check_in"]:
            wh    = "Checkout Missing"
        else:
            wh    = "0 hrs"

        result.append({"id": r["id"], "name": r["person_name"], "status": r["status"], "working_hours": wh})

    cursor.close(); conn.close()
    return result


@app.get("/yearly-attendance")
def generate_year_excel(year: int = 2026):
    conn   = get_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT DISTINCT person_name FROM attendance ORDER BY person_name")
    employees = cursor.fetchall()

    wb = Workbook()
    wb.remove(wb.active)
    today = datetime.today().date()

    for month in range(1, 13):
        days_in_month = calendar.monthrange(year, month)[1]
        ws = wb.create_sheet(title=calendar.month_name[month][:3].upper())

        headers = ["S.No", "Name"] + [str(d) for d in range(1, days_in_month + 1)] + MONTH_END_COLUMNS
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            if col == 1:
                cell.fill = sno_bg_fill
                cell.font = Font(bold=True, color="2C3E6B", name="Calibri", size=11)
            elif col == 2:
                cell.fill = name_bg_fill
                cell.font = header_font
            elif col >= (3 + days_in_month):
                idx = col - (3 + days_in_month)
                cell.fill = month_end_fills[idx]
                cell.font = Font(bold=True, color="1A2E5A", name="Calibri", size=10)
            else:
                cell.fill = PatternFill(start_color="EEF1F8", end_color="EEF1F8", fill_type="solid")
                cell.font = Font(bold=True, color="3A4A7A", name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = header_border

        max_working_days = sum(
            1 for d in range(1, days_in_month + 1)
            if (datetime(year, month, d).date() <= today)
            and not _is_off_day(year, month, d)
            and f"{year}-{month:02d}-{d:02d}" not in FESTIVAL_DAYS
        )

        for emp_idx, emp in enumerate(employees, start=1):
            name = emp["person_name"]
            morning_count     = 0
            present_count     = 0
            leave_count       = 0
            db_absent_count   = 0
            checkout_missing  = 0
            total_checkout    = 0
            in_times          = []
            row_data          = [emp_idx, name]
            cell_data         = []

            def _fmt_time(t):
                """Format a timedelta or time object as HH:MM:SS string."""
                if t is None:
                    return None
                if hasattr(t, "total_seconds"):   # timedelta from MySQL TIME col
                    total = int(t.total_seconds())
                    h, rem = divmod(abs(total), 3600)
                    m, s   = divmod(rem, 60)
                    return f"{h:02d}:{m:02d}:{s:02d}"
                return t.strftime("%H:%M:%S")      # datetime.time

            def _duration(ci, co):
                """Return human-readable duration string, e.g. '5h 19m'."""
                if ci is None or co is None:
                    return ""
                def _secs(t):
                    if hasattr(t, "total_seconds"):
                        return int(t.total_seconds())
                    return t.hour * 3600 + t.minute * 60 + t.second
                diff = _secs(co) - _secs(ci)
                if diff < 0:
                    diff = 0
                h, rem = divmod(diff, 3600)
                m      = rem // 60
                return f"{h}h {m}m"

            for d in range(1, days_in_month + 1):
                date_str   = f"{year}-{month:02d}-{d:02d}"
                date_obj   = datetime(year, month, d).date()
                is_future  = date_obj > today
                is_off     = _is_off_day(year, month, d)
                is_festival = date_str in FESTIVAL_DAYS
                is_working  = not is_off and not is_festival and not is_future

                cursor.execute(
                    "SELECT check_in, check_out, status FROM attendance WHERE person_name=%s AND date=%s",
                    (name, date_str),
                )
                record = cursor.fetchone()
                info   = {"text": "", "fill": None, "type": "blank"}

                if record:
                    ci, co, st = record["check_in"], record["check_out"], record["status"]
                    ci_str = _fmt_time(ci)
                    co_str = _fmt_time(co)
                    if is_off and not is_festival:
                        # Worked on a day-off — show times with P* label kept compact
                        if ci_str and co_str:
                            dur  = _duration(ci, co)
                            text = f"{ci_str}\n{co_str}\n{dur}"
                        elif ci_str:
                            text = f"{ci_str}\nCHECKOUT\nMISSING"
                        else:
                            text = "WEEK OFF"
                        info = {"text": text, "fill": present_fill, "type": "present"}
                    elif ci_str and co_str:
                        dur = _duration(ci, co)
                        text = f"{ci_str}\n{co_str}\n{dur}"
                        in_times.append(ci)
                        morning_count  += 1
                        total_checkout += 1
                        present_count  += 1
                        info = {"text": text, "fill": present_fill, "type": "present"}
                    elif ci_str and not co_str:
                        text = f"{ci_str}\nCHECKOUT\nMISSING"
                        in_times.append(ci)
                        morning_count    += 1
                        checkout_missing += 1
                        info = {"text": text, "fill": checkout_missing_fill, "type": "checkout_missing"}
                elif is_future:
                    info = {"text": "", "fill": None, "type": "blank"}
                elif is_off:
                    if is_festival:
                        info = {"text": "FESTIVAL", "fill": festival_fill,  "type": "festival"}
                    else:
                        info = {"text": "WEEK OFF", "fill": sunday_fill,    "type": "sunday"}
                elif is_festival:
                    info = {"text": "FESTIVAL", "fill": festival_fill, "type": "festival"}
                elif is_working:
                    info = {"text": "ABSENT",   "fill": absent_fill,   "type": "absent"}
                    db_absent_count += 1

                cell_data.append(info)
                row_data.append(info["text"])

            paid_leave_month    = 1
            actual_leave_taken  = db_absent_count + leave_count
            # Deduction = absences not covered by paid leave allowance (never negative)
            actual_deduction    = max(0, actual_leave_taken - paid_leave_month)
            # Unused paid leave remaining
            paid_leave_balance  = max(0, paid_leave_month - leave_count - db_absent_count)
            carry_fwd           = 1 if (leave_count + db_absent_count + checkout_missing) == 0 else 0
            # Payable days = days actually present + absences covered by leave
            total_days          = present_count + min(actual_leave_taken, paid_leave_month)
            in_avg              = "-"
            if in_times:
                def _to_minutes(t):
                    if hasattr(t, "total_seconds"):
                        return int(t.total_seconds()) // 60
                    return t.hour * 60 + t.minute
                avg_min = sum(_to_minutes(t) for t in in_times) // len(in_times)
                in_avg  = f"{avg_min//60:02d}:{avg_min%60:02d}"

            row_data += [paid_leave_month, actual_leave_taken, actual_deduction,
                         paid_leave_balance, carry_fwd, total_days, in_avg,
                         morning_count, total_checkout]

            ws.append(row_data)
            _apply_excel_styling(ws, ws.max_row, cell_data, days_in_month, name)

        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 28
        for c in range(3, 3 + days_in_month):
            ws.column_dimensions[get_column_letter(c)].width = 14   # wider for timestamps
        ws.row_dimensions[1].height = 30
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 52   # taller rows for 3-line timestamp cells
        for i, w in enumerate([22, 18, 16, 20, 26, 12, 14, 12, 12]):
            ws.column_dimensions[get_column_letter(3 + days_in_month + i)].width = w

    fname = f"attendance_{year}_complete.xlsx"
    wb.save(fname)
    cursor.close(); conn.close()
    return FileResponse(fname, filename=fname)



@app.get("/debug-db")
def debug_db():
    """Inspect raw embedding dims in the faces table — helps diagnose FAISS assertion errors."""
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, embedding FROM faces")
    rows   = cursor.fetchall()
    cursor.close(); conn.close()
    result = []
    for rid, name, blob in rows:
        emb  = np.frombuffer(blob, dtype=np.float32)
        result.append({"id": rid, "name": name, "embedding_dim": int(emb.shape[0]),
                        "expected_dim": FAISS_DIM, "ok": emb.shape[0] == FAISS_DIM})
    bad = [r for r in result if not r["ok"]]
    return {
        "total_rows": len(result),
        "bad_rows": len(bad),
        "faiss_index_count": faiss_index.ntotal,
        "note": "Run POST /build-db to purge bad rows and rebuild from dataset images.",
        "rows": result,
    }

# ─────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)